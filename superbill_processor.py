"""
Superbill Processor
-------------------
UI tool to:
 - Search for an input Superbill XLS/XLSX file by name
 - Select an output XLS/XLSX file
 - Validate and drop expected empty columns
 - Append new rows (with column remapping) to the output file
 - Skip duplicate rows (identified by cols A, B, O in input / A, C, I in output)
"""

import sys
import tkinter as tk
from tkinter import ttk, filedialog
import tkinterdnd2 as dnd
import re
import os
import shutil
import tempfile
import threading
import traceback
from datetime import datetime

import pandas as pd
import openpyxl
from dotenv import load_dotenv, set_key
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

from lynx_flow import download_superbill, month_date_range

# ── Constants ────────────────────────────────────────────────────────────────

# Header row inside the Superbill file (0-based row index)
INPUT_HEADER_ROW = 6

# Expected completely-empty column names in the input file.
# The raw Excel headers contain newlines, so we normalise by replacing \n with space.
EXPECTED_EMPTY_COLS_NORMALISED = {
    "Primary Carrier",
    "Primary Policy",
    "Secondary Carrier",
    "Secondary Policy",
    "Tertiary Carrier",
    "Tertiary Policy",
    "Clinical Trial",
    "Seq No",
    "Comment",
}

# Input column indices (0-based, after reading with header=INPUT_HEADER_ROW)
# A=0 B=1 C=2 D=3 E=4 F=5 G=6 H=7 O=14 P=15 Q=16 R=17 S=18 T=19 U=20 V=21
# AD=29 AE=30 AF=31
COL_MAP_INPUT_TO_OUTPUT = {
    0:  0,   # A  → A
    1:  2,   # B  → C
    2:  3,   # C  → D
    3:  4,   # D  → E
    4:  5,   # E  → F
    5:  6,   # F  → G
    6:  7,   # G  → H
    14: 8,   # O  → I
    15: 9,   # P  → J
    16: 10,  # Q  → K
    17: 11,  # R  → L
    18: 12,  # S  → M
    19: 13,  # T  → N
    20: 14,  # U  → O
    21: 15,  # V  → P
    29: 23,  # AD → X
    30: 24,  # AE → Y
    31: 25,  # AF → Z
}

# Identity: input col indices [A, B, O] and output col indices [A, C, I]
IDENTITY_INPUT_COLS  = [0, 1, 14]   # A, B, O
IDENTITY_OUTPUT_COLS = [0, 2, 8]    # A, C, I

# ── Helpers ──────────────────────────────────────────────────────────────────

def col_letter(idx):
    """Return Excel-style column letter for 0-based index."""
    result = ""
    n = idx + 1
    while n:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


def normalise_col(name):
    """Replace newlines/extra spaces in column name."""
    if not isinstance(name, str):
        return str(name)
    return " ".join(name.split())


# ── Core processing ──────────────────────────────────────────────────────────

def _last_nonempty_row(ws):
    """Return the 1-based index of the last row that has at least one non-None cell."""
    last = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if any(v is not None for v in row):
            last = i
    return last


def process(input_path, output_path, log, confirm):
    """
    Main processing function.
      log(msg)        – append a line to the UI message pane
      confirm(msg)    – show a Proceed/Abort prompt; blocks until user responds;
                        returns True (proceed) or False (abort)
    Returns True on success/skip, False on abort/error.
    """

    # ── 1. Read input ────────────────────────────────────────────────────────
    log(f"Reading input: {os.path.basename(input_path)}")
    try:
        df_in = pd.read_excel(input_path, header=INPUT_HEADER_ROW)
    except Exception as e:
        log(f"ERROR reading input file: {e}")
        return False

    df_in.columns = [normalise_col(c) for c in df_in.columns]
    df_in.dropna(how="all", inplace=True)
    df_in.reset_index(drop=True, inplace=True)
    log(f"  Input rows (after cleanup): {len(df_in)}  |  Columns: {len(df_in.columns)}")

    # ── 2. Detect empty columns & compare with expected list ─────────────────
    actual_empty = {c for c in df_in.columns if df_in[c].isna().all()}
    missing_from_actual = EXPECTED_EMPTY_COLS_NORMALISED - actual_empty
    extra_actual_empty  = actual_empty - EXPECTED_EMPTY_COLS_NORMALISED

    if missing_from_actual or extra_actual_empty:
        log("⚠  DISCREPANCY between expected empty columns and actual empty columns:")
        if missing_from_actual:
            log("   Expected to be empty but are NOT empty:")
            for c in sorted(missing_from_actual):
                log(f"     • {c}")
        if extra_actual_empty:
            log("   Empty but NOT in the expected list:")
            for c in sorted(extra_actual_empty):
                log(f"     • {c}")
        log("   (Continuing – actually-empty columns will be removed.)")
    else:
        log("✓  Empty columns match expected list exactly.")

    # Reload original with all columns for index-based mapping
    try:
        df_raw = pd.read_excel(input_path, header=INPUT_HEADER_ROW)
    except Exception as e:
        log(f"ERROR re-reading input file: {e}")
        return False
    df_raw.columns = [normalise_col(c) for c in df_raw.columns]
    df_raw.dropna(how="all", inplace=True)
    df_raw.reset_index(drop=True, inplace=True)

    # Merged cells in Excel produce NaN in all but the first row of the merged range.
    # Forward-fill identity columns so every sub-row inherits the parent cell's value.
    for _col_idx in IDENTITY_INPUT_COLS:
        if _col_idx < len(df_raw.columns):
            df_raw.iloc[:, _col_idx] = df_raw.iloc[:, _col_idx].ffill()

    raw_cols = list(df_raw.columns)

    # ── 3. Load output file ───────────────────────────────────────────────────
    if not os.path.isfile(output_path):
        log("ERROR: output file does not exist. Please select an existing output file.")
        return False

    log(f"Loading output file: {os.path.basename(output_path)}")
    try:
        wb_out = load_workbook(output_path)
        ws_out = wb_out.active
    except Exception as e:
        log(f"ERROR reading output file: {e}")
        return False

    out_rows = list(ws_out.iter_rows(values_only=True))

    # Build identity key set from existing output rows (skip header row 0)
    existing_keys = set()
    for r in out_rows[1:]:
        def _val(row, idx):
            v = row[idx] if idx < len(row) else None
            return str(v).strip() if v is not None else ""
        key = (
            _val(r, IDENTITY_OUTPUT_COLS[0]),
            _val(r, IDENTITY_OUTPUT_COLS[1]),
            _val(r, IDENTITY_OUTPUT_COLS[2]),
        )
        existing_keys.add(key)

    # ── 4. Duplicate detection ────────────────────────────────────────────────
    duplicates   = []
    rows_to_add  = []

    for idx, row in df_raw.iterrows():
        def cell(col_idx, _row=row):
            if col_idx >= len(raw_cols):
                return ""
            v = _row.iloc[col_idx]
            return str(v).strip() if pd.notna(v) else ""

        key = (
            cell(IDENTITY_INPUT_COLS[0]),
            cell(IDENTITY_INPUT_COLS[1]),
            cell(IDENTITY_INPUT_COLS[2]),
        )
        # Skip rows with no identifying information
        if all(v == "" for v in key):
            continue
        excel_row = INPUT_HEADER_ROW + 2 + idx  # 1-based Excel row number
        if key in existing_keys:
            duplicates.append((excel_row, key))
        else:
            rows_to_add.append(row)

    if duplicates:
        log("")
        log(f"⚠  {len(duplicates)} duplicate row(s) already exist in the output:")
        log(f"    {'Row':<5}  {'Date of Service':<15}  {'Patient Name':<30}  {'Billing Code'}")
        log(f"    {'-'*5}  {'-'*15}  {'-'*30}  {'-'*15}")
        for excel_row, d in duplicates:
            log(f"    {excel_row:<5}  {d[0]:<15}  {d[1]:<30}  {d[2]}")
        log("")
        log(f"  {len(rows_to_add)} new row(s) would be appended.")
        if not confirm(f"{len(duplicates)} duplicate(s) found (listed above).\n"
                       f"Proceed with appending {len(rows_to_add)} new row(s)?"):
            log("⛔  Aborted by user. Output file was NOT modified.")
            return False
    else:
        log(f"✓  No duplicates found. {len(rows_to_add)} row(s) will be appended.")

    if not rows_to_add:
        log("ℹ  Nothing new to add – all input rows are already in the output.")
        return True

    # ── 5. Backup output file before writing ──────────────────────────────────
    backup_fd, backup_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(backup_fd)
    try:
        shutil.copy2(output_path, backup_path)
    except Exception as e:
        log(f"ERROR creating backup: {e}")
        return False

    # ── 6. Find last non-empty row and append after it ────────────────────────
    insert_after = _last_nonempty_row(ws_out)
    # Truncate trailing empty rows only when we have a non-empty anchor row.
    # For a brand-new/empty workbook insert_after can be 0 while max_row is 1,
    # and deleting rows in a loop may never converge.
    if insert_after > 0:
        while ws_out.max_row > insert_after:
            ws_out.delete_rows(ws_out.max_row)
    else:
        log("  Output workbook appears empty; appending will start at the first row.")

    log(f"  Appending after row {insert_after} (last non-empty row in output).")

    max_out_col = max(COL_MAP_INPUT_TO_OUTPUT.values()) + 1
    appended_keys = []
    for row in rows_to_add:
        out_row = [None] * max_out_col
        for in_idx, out_idx in COL_MAP_INPUT_TO_OUTPUT.items():
            if in_idx < len(raw_cols):
                v = row.iloc[in_idx]
                out_row[out_idx] = None if pd.isna(v) else v
        ws_out.append(out_row)
        # Track what we added by identity key
        def _cell(col_idx, _row=row):
            if col_idx >= len(raw_cols):
                return ""
            v = _row.iloc[col_idx]
            return str(v).strip() if pd.notna(v) else ""
        appended_keys.append((
            _cell(IDENTITY_INPUT_COLS[0]),
            _cell(IDENTITY_INPUT_COLS[1]),
            _cell(IDENTITY_INPUT_COLS[2]),
        ))

    # ── 7. Save to output file ────────────────────────────────────────────────
    try:
        wb_out.save(output_path)
    except Exception as e:
        log(f"ERROR saving output file: {e}")
        log("  Restoring backup…")
        shutil.copy2(backup_path, output_path)
        os.unlink(backup_path)
        return False

    log(f"  Saved. Verifying written data…")

    # ── 8. Verify: re-read output and confirm all appended rows are present ───
    try:
        wb_verify = load_workbook(output_path)
        ws_verify = wb_verify.active
        verified_keys = set()
        for r in ws_verify.iter_rows(values_only=True):
            def _vv(row, idx):
                v = row[idx] if idx < len(row) else None
                return str(v).strip() if v is not None else ""
            verified_keys.add((
                _vv(r, IDENTITY_OUTPUT_COLS[0]),
                _vv(r, IDENTITY_OUTPUT_COLS[1]),
                _vv(r, IDENTITY_OUTPUT_COLS[2]),
            ))
    except Exception as e:
        log(f"ERROR during verification read: {e}")
        verified_keys = set()

    missing_after_write = [k for k in appended_keys if k not in verified_keys]

    if missing_after_write:
        log("")
        log(f"⚠  VERIFICATION FAILED: {len(missing_after_write)} row(s) not found in output after save:")
        log(f"    {'Date of Service':<15}  {'Patient Name':<30}  {'Billing Code'}")
        log(f"    {'-'*15}  {'-'*30}  {'-'*15}")
        for k in missing_after_write:
            log(f"    {k[0]:<15}  {k[1]:<30}  {k[2]}")
        log("")
        if not confirm(f"Verification found {len(missing_after_write)} missing row(s).\n"
                       "Keep the (possibly incomplete) output, or Abort to restore the original?"):
            log("⛔  Aborted – restoring original output file from backup.")
            shutil.copy2(backup_path, output_path)
            os.unlink(backup_path)
            return False
        else:
            log("⚠  User chose to keep output despite verification discrepancy.")
    else:
        log(f"✓  Verification passed – all {len(appended_keys)} row(s) confirmed in output.")

    os.unlink(backup_path)
    log("")
    log(f"✅  Done.  {len(appended_keys)} row(s) added to {os.path.basename(output_path)}.")
    return True


# ── UI ────────────────────────────────────────────────────────────────────────

_README_PATH = os.path.join(
    getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__))),
    "README.md"
)
_ENV_PATH = os.path.join(
    getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__))),
    ".env"
)


def _show_about():
    """Open a Toplevel window that renders README.md with basic Markdown styling."""
    win = tk.Toplevel()
    win.title("About – Superbill Processor")
    win.geometry("720x580")
    win.resizable(True, True)

    frm = ttk.Frame(win)
    frm.pack(fill="both", expand=True, padx=10, pady=10)

    txt = tk.Text(frm, wrap="word", state="disabled",
                  font=("Segoe UI", 10), background="#ffffff",
                  foreground="#1e1e1e", relief="flat", padx=12, pady=8)
    txt.pack(side="left", fill="both", expand=True)
    sb = ttk.Scrollbar(frm, orient="vertical", command=txt.yview)
    sb.pack(side="right", fill="y")
    txt.configure(yscrollcommand=sb.set)

    # Define tags
    txt.tag_configure("h1",    font=("Segoe UI", 18, "bold"), spacing3=6)
    txt.tag_configure("h2",    font=("Segoe UI", 14, "bold"), spacing1=10, spacing3=4)
    txt.tag_configure("h3",    font=("Segoe UI", 11, "bold"), spacing1=8, spacing3=2)
    txt.tag_configure("bold",  font=("Segoe UI", 10, "bold"))
    txt.tag_configure("code",  font=("Consolas", 9), background="#f0f0f0", foreground="#c7254e")
    txt.tag_configure("block", font=("Consolas", 9), background="#f5f5f5",
                      lmargin1=20, lmargin2=20, spacing1=2, spacing3=2)
    txt.tag_configure("bullet", lmargin1=16, lmargin2=28)
    txt.tag_configure("table",  font=("Consolas", 9), lmargin1=16, lmargin2=16)
    txt.tag_configure("normal", font=("Segoe UI", 10))

    try:
        with open(_README_PATH, encoding="utf-8") as f:
            lines = f.readlines()
    except FileNotFoundError:
        lines = ["*README.md not found.*\n"]

    txt.configure(state="normal")
    in_code_block = False
    for raw in lines:
        line = raw.rstrip("\n")

        # Fenced code block toggle
        if line.startswith("```"):
            in_code_block = not in_code_block
            txt.insert("end", "\n")
            continue

        if in_code_block:
            txt.insert("end", line + "\n", "block")
            continue

        # Table rows (contain |)
        if "|" in line and line.strip().startswith("|"):
            # Skip separator rows like |---|---|
            if re.fullmatch(r"[\s|\-:]+", line):
                continue
            txt.insert("end", line + "\n", "table")
            continue

        # Headings
        m = re.match(r"^(#{1,3})\s+(.*)", line)
        if m:
            level = len(m.group(1))
            txt.insert("end", m.group(2) + "\n", ("h1", "h2", "h3")[level - 1])
            continue

        # Bullet list items
        m = re.match(r"^[-*]\s+(.*)", line)
        if m:
            _insert_inline(txt, "• " + m.group(1) + "\n", "bullet")
            continue

        # Blank line
        if not line.strip():
            txt.insert("end", "\n")
            continue

        # Normal paragraph line
        _insert_inline(txt, line + "\n", "normal")

    txt.configure(state="disabled")
    ttk.Button(win, text="Close", command=win.destroy).pack(pady=(0, 8))


def _insert_inline(txt: tk.Text, text: str, base_tag: str):
    """Insert text into a tk.Text widget, handling **bold** and `inline code`."""
    pattern = re.compile(r"(\*\*.+?\*\*|`.+?`)")   
    pos = 0
    for m in pattern.finditer(text):
        # plain segment before the match
        if m.start() > pos:
            txt.insert("end", text[pos:m.start()], base_tag)
        token = m.group()
        if token.startswith("**"):
            txt.insert("end", token[2:-2], (base_tag, "bold"))
        else:  # backtick
            txt.insert("end", token[1:-1], (base_tag, "code"))
        pos = m.end()
    if pos < len(text):
        txt.insert("end", text[pos:], base_tag)


class ToolTip:
    def __init__(self, widget, text: str):
        self.widget = widget
        self.text = text
        self.tipwindow = None
        self.widget.bind("<Enter>", self._show)
        self.widget.bind("<Leave>", self._hide)

    def _show(self, _event=None):
        if self.tipwindow or not self.text:
            return
        x = self.widget.winfo_rootx() + 18
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 6
        self.tipwindow = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        tk.Label(
            tw,
            text=self.text,
            justify="left",
            background="#ffffe0",
            relief="solid",
            borderwidth=1,
            padx=6,
            pady=4,
            font=("Segoe UI", 9),
        ).pack()

    def _hide(self, _event=None):
        if self.tipwindow is not None:
            self.tipwindow.destroy()
            self.tipwindow = None


class App(dnd.TkinterDnD.Tk):
    def __init__(self):
        super().__init__()
        self.title("Superbill Processor")
        self.resizable(True, True)
        self.minsize(780, 620)
        load_dotenv(_ENV_PATH, override=True)
        self._confirm_event = threading.Event()
        self._confirm_result = False
        self._fetch_running = False
        self._merge_running = False
        self._merge_abort_requested = False
        self._fetch_abort_requested = False
        self._fetch_browser = None
        self._fetch_context = None
        self._build_ui()

    # ── layout ────────────────────────────────────────────────────────────────

    def _build_ui(self):
        pad = dict(padx=10, pady=4)

        # ── Fetch section ──────────────────────────────────────────────────────
        frm_fetch = ttk.LabelFrame(self, text="Lynx Fetch")
        frm_fetch.pack(fill="x", **pad)

        now = datetime.now()
        self.fetch_month_num_var = tk.IntVar(value=now.month)
        self.fetch_year_var = tk.IntVar(value=now.year)
        self.fetch_slow_seconds_var = tk.IntVar(value=2)
        self.fetch_headless_var = tk.BooleanVar(value=False)
        self.fetch_download_dir_var = tk.StringVar(value=self._env_get("LYNX_WORK_FOLDER", "data"))
        self.fetch_result_name_var = tk.StringVar(value="Downloaded file: (none)")
        self.fetch_status_var = tk.StringVar(value="Fetch status: idle")

        # Keep settings at the top of the Lynx Fetch section
        settings_row = ttk.Frame(frm_fetch)
        settings_row.grid(row=0, column=0, columnspan=5, sticky="w", padx=6, pady=(4, 2))
        ttk.Button(settings_row, text="⚙  Settings", command=self._open_fetch_settings, width=12).pack(side="left")
        ttk.Button(settings_row, text="ℹ  About", command=_show_about, width=12).pack(side="left", padx=6)

        ttk.Label(frm_fetch, text="Month/Year:").grid(row=1, column=0, sticky="w", padx=6, pady=4)
        month_year = ttk.Frame(frm_fetch)
        month_year.grid(row=1, column=1, sticky="w", padx=4, pady=4)
        ttk.Spinbox(month_year, from_=1, to=12, textvariable=self.fetch_month_num_var, width=4, format="%02.0f").pack(
            side="left"
        )
        ttk.Label(month_year, text="/").pack(side="left", padx=4)
        ttk.Spinbox(month_year, from_=2000, to=2100, textvariable=self.fetch_year_var, width=6).pack(side="left")
        ttk.Label(frm_fetch, text="Slow-mo (sec):").grid(row=1, column=2, sticky="w", padx=16, pady=4)
        ttk.Spinbox(frm_fetch, from_=0, to=30, textvariable=self.fetch_slow_seconds_var, width=8).grid(
            row=1, column=3, sticky="w", padx=4, pady=4
        )
        ttk.Checkbutton(frm_fetch, text="Headless", variable=self.fetch_headless_var).grid(
            row=1, column=4, sticky="w", padx=16, pady=4
        )

        ttk.Label(frm_fetch, text="Download folder:").grid(row=2, column=0, sticky="w", padx=6, pady=4)
        ttk.Entry(frm_fetch, textvariable=self.fetch_download_dir_var, width=54).grid(
            row=2, column=1, columnspan=3, sticky="ew", padx=4, pady=4
        )
        ttk.Button(frm_fetch, text="Browse…", command=self._browse_fetch_download_dir).grid(
            row=2, column=4, sticky="e", padx=4, pady=4
        )

        fetch_btn_row = ttk.Frame(frm_fetch)
        fetch_btn_row.grid(row=3, column=0, columnspan=5, sticky="w", padx=6, pady=(4, 6))
        self.fetch_btn = tk.Button(fetch_btn_row, text="⇩  Run Fetch", command=self._run_fetch, width=16, height=2)
        self.fetch_btn._base_label = "Run Fetch"
        self.fetch_btn.pack(side="left")
        self._set_action_button_state(self.fetch_btn, "idle")
        ToolTip(
            self.fetch_btn,
            "Run Fetch button colors: blue=idle/running, green=last run succeeded, red=last run failed.",
        )

        ttk.Label(frm_fetch, textvariable=self.fetch_result_name_var).grid(
            row=4, column=0, columnspan=5, sticky="w", padx=8, pady=(0, 2)
        )
        ttk.Label(frm_fetch, textvariable=self.fetch_status_var).grid(
            row=5, column=0, columnspan=5, sticky="w", padx=8, pady=(0, 4)
        )
        frm_fetch.columnconfigure(3, weight=1)

        # ── Input file section ────────────────────────────────────────────────
        frm_in = ttk.LabelFrame(self, text="Input Superbill file")
        frm_in.pack(fill="x", **pad)

        self.input_path_var = tk.StringVar()
        ttk.Label(frm_in, text="File:").grid(row=0, column=0, sticky="w", padx=6, pady=4)
        ent_in = ttk.Entry(frm_in, textvariable=self.input_path_var, width=55)
        ent_in.grid(row=0, column=1, sticky="ew", padx=4, pady=4)
        ttk.Button(frm_in, text="Browse…", command=self._browse_input).grid(row=0, column=2, padx=4)
        frm_in.columnconfigure(1, weight=1)
        ent_in.drop_target_register(dnd.DND_FILES)
        ent_in.dnd_bind("<<Drop>>", lambda e: self.input_path_var.set(self._clean_drop(e.data)))

        # ── Output file section ───────────────────────────────────────────────
        frm_out = ttk.LabelFrame(self, text="Output file")
        frm_out.pack(fill="x", **pad)

        self.output_path_var = tk.StringVar()
        default_output = self._env_get("OUTPUT_FILE_PATH", "")
        if default_output:
            self.output_path_var.set(default_output)
        ttk.Label(frm_out, text="File:").grid(row=0, column=0, sticky="w", padx=6, pady=4)
        ent_out = ttk.Entry(frm_out, textvariable=self.output_path_var, width=55)
        ent_out.grid(row=0, column=1, sticky="ew", padx=4, pady=4)
        ttk.Button(frm_out, text="Open…", command=self._browse_output).grid(
            row=0, column=2, padx=4)
        frm_out.columnconfigure(1, weight=1)
        ent_out.drop_target_register(dnd.DND_FILES)
        ent_out.dnd_bind("<<Drop>>", lambda e: self.output_path_var.set(self._clean_drop(e.data)))

        # ── Run button ────────────────────────────────────────────────────────
        btn_row_top = ttk.Frame(self)
        btn_row_top.pack(pady=6, fill="x", padx=10)
        self.run_btn = tk.Button(btn_row_top, text="▶  Run Merge", command=self._run, width=20, height=2)
        self.run_btn._base_label = "Run Merge"
        self.run_btn.pack(side="left")
        self._set_action_button_state(self.run_btn, "idle")
        ToolTip(
            self.run_btn,
            "Run Merge button colors: blue=idle/running, green=last run succeeded, red=last run failed.",
        )
        self.merge_status_var = tk.StringVar(value="Merge status: idle")
        ttk.Label(btn_row_top, textvariable=self.merge_status_var).pack(side="left", padx=12)
        self.lifecycle_btn = tk.Button(btn_row_top, command=self._lifecycle_action, width=16, height=2)
        self.lifecycle_btn.pack(side="right")
        ToolTip(
            self.lifecycle_btn,
            "Lifecycle button: yellow Cancel before start, red Abort while running, green Close after success.",
        )
        self._set_lifecycle_button_state("cancel")
        self.input_path_var.trace_add("write", lambda *_: self._refresh_merge_button_enabled())
        self.output_path_var.trace_add("write", lambda *_: self._refresh_merge_button_enabled())
        self._refresh_merge_button_enabled()

        # ── Confirm / Abort prompt (hidden until needed) ──────────────────────
        self.frm_confirm = ttk.LabelFrame(self, text="Confirmation required")
        # not packed yet – shown dynamically

        self.confirm_msg_var = tk.StringVar()
        ttk.Label(self.frm_confirm, textvariable=self.confirm_msg_var,
                  foreground="#cc8800", wraplength=640,
                  justify="left").pack(padx=8, pady=6, anchor="w")

        btn_row = ttk.Frame(self.frm_confirm)
        btn_row.pack(pady=(0, 8))
        ttk.Button(btn_row, text="✔  Proceed",
                   command=lambda: self._resolve_confirm(True),
                   width=16).pack(side="left", padx=12)
        ttk.Button(btn_row, text="✘  Abort",
                   command=lambda: self._resolve_confirm(False),
                   width=16).pack(side="left", padx=12)

        # ── Messages pane ─────────────────────────────────────────────────────
        frm_log = ttk.LabelFrame(self, text="Messages")
        frm_log.pack(fill="both", expand=True, **pad)

        self.log_text = tk.Text(frm_log, state="disabled", wrap="word",
                                font=("Consolas", 9), background="#1e1e1e",
                                foreground="#d4d4d4", insertbackground="white")
        self.log_text.pack(side="left", fill="both", expand=True)
        sb2 = ttk.Scrollbar(frm_log, orient="vertical", command=self.log_text.yview)
        sb2.pack(side="right", fill="y")
        self.log_text.configure(yscrollcommand=sb2.set)

        ttk.Button(self, text="Clear log", command=self._clear_log).pack(
            anchor="e", padx=10, pady=(0, 6))

    @staticmethod
    def _env_get(key: str, default: str = "") -> str:
        return (os.environ.get(key) or default).strip()

    @staticmethod
    def _month_text(month_num: int, year: int) -> str:
        return f"{int(month_num):02d}/{int(year)}"

    @staticmethod
    def _set_action_button_state(btn: tk.Button, state_name: str):
        palette = {
            "idle": "#1f6feb",
            "running": "#1f6feb",
            "success": "#2ea043",
            "failed": "#d73a49",
        }
        c = palette.get(state_name, palette["idle"])
        base_label = getattr(btn, "_base_label", btn.cget("text"))
        icon = {
            "idle": "▶",
            "running": "▶",
            "success": "✔",
            "failed": "✘",
        }.get(state_name, "▶")
        btn.configure(
            text=f"{icon}  {base_label}",
            bg=c,
            activebackground=c,
            fg="white",
            activeforeground="white",
            disabledforeground="white",
            relief="raised",
            bd=1,
        )

    def _set_lifecycle_button_state(self, state_name: str):
        # cancel (yellow), abort (red), close (green)
        if state_name == "abort":
            self.lifecycle_btn.configure(
                text="✘  Abort",
                bg="#d73a49",
                activebackground="#d73a49",
                fg="white",
                activeforeground="white",
            )
        elif state_name == "close":
            self.lifecycle_btn.configure(
                text="✔  Close",
                bg="#2ea043",
                activebackground="#2ea043",
                fg="white",
                activeforeground="white",
            )
        else:
            self.lifecycle_btn.configure(
                text="⚠  Cancel",
                bg="#d4a72c",
                activebackground="#d4a72c",
                fg="black",
                activeforeground="black",
            )

    def _refresh_lifecycle_state(self):
        if self._fetch_running or self._merge_running:
            self._set_lifecycle_button_state("abort")
            return
        # Once a run succeeded, present Close; otherwise keep Cancel
        fetch_ok = "finished successfully" in self.fetch_status_var.get().lower()
        merge_ok = "finished successfully" in self.merge_status_var.get().lower()
        if fetch_ok or merge_ok:
            self._set_lifecycle_button_state("close")
        else:
            self._set_lifecycle_button_state("cancel")

    def _lifecycle_action(self):
        if self._fetch_running or self._merge_running:
            self._log("Abort requested by user.")
            self._fetch_abort_requested = True
            self._merge_abort_requested = True
            # If merge waits for confirmation, resolve to abort now.
            if hasattr(self, "_confirm_event") and not self._confirm_event.is_set():
                self._resolve_confirm(False)
            # Try to stop Playwright fetch quickly.
            try:
                if self._fetch_context is not None:
                    self._fetch_context.close()
            except Exception:
                pass
            try:
                if self._fetch_browser is not None:
                    self._fetch_browser.close()
            except Exception:
                pass
            return
        # Not running: Cancel/Close both close the app.
        self.destroy()

    @staticmethod
    def _clean_drop(data: str) -> str:
        """Normalize a drag-and-drop path: strip curly braces added for paths with spaces."""
        data = data.strip()
        if data.startswith("{") and data.endswith("}"):
            data = data[1:-1]
        return data

    def _merge_preconditions_ok(self) -> bool:
        input_path = self.input_path_var.get().strip()
        output_path = self.output_path_var.get().strip()
        return bool(input_path and output_path and os.path.isfile(input_path) and os.path.isfile(output_path))

    def _refresh_merge_button_enabled(self):
        if self._merge_running:
            self.run_btn.configure(state="disabled")
            return
        self.run_btn.configure(state=("normal" if self._merge_preconditions_ok() else "disabled"))

    def _browse_input(self):
        path = filedialog.askopenfilename(
            title="Select input Superbill file",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")])
        if path:
            self.input_path_var.set(path)
            self._refresh_merge_button_enabled()

    def _browse_output(self):
        path = filedialog.askopenfilename(
            title="Select existing output file to append to",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")])
        if path:
            self.output_path_var.set(path)
            self._refresh_merge_button_enabled()

    def _browse_fetch_download_dir(self):
        path = filedialog.askdirectory(title="Select download folder")
        if path:
            self.fetch_download_dir_var.set(path)

    def _open_fetch_settings(self):
        win = tk.Toplevel(self)
        win.title("Fetch Settings")
        win.resizable(False, False)
        win.transient(self)
        win.grab_set()

        url_var = tk.StringVar(value=self._env_get("LYNX_URL"))
        user_var = tk.StringVar(value=self._env_get("LYNX_USER"))
        pw_var = tk.StringVar(value=self._env_get("LYNX_PASSWORD"))
        work_var = tk.StringVar(value=self._env_get("LYNX_WORK_FOLDER", "data"))
        output_var = tk.StringVar(value=self._env_get("OUTPUT_FILE_PATH", self.output_path_var.get().strip()))

        frm = ttk.Frame(win, padding=10)
        frm.pack(fill="both", expand=True)

        ttk.Label(frm, text="Lynx URL").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        ttk.Entry(frm, textvariable=url_var, width=58).grid(row=0, column=1, sticky="ew", padx=4, pady=4)
        ttk.Label(frm, text="Lynx User").grid(row=1, column=0, sticky="w", padx=4, pady=4)
        ttk.Entry(frm, textvariable=user_var, width=58).grid(row=1, column=1, sticky="ew", padx=4, pady=4)
        ttk.Label(frm, text="Lynx Password").grid(row=2, column=0, sticky="w", padx=4, pady=4)
        ttk.Entry(frm, textvariable=pw_var, width=58, show="*").grid(row=2, column=1, sticky="ew", padx=4, pady=4)
        ttk.Label(frm, text="Work Folder").grid(row=3, column=0, sticky="w", padx=4, pady=4)
        ttk.Entry(frm, textvariable=work_var, width=58).grid(row=3, column=1, sticky="ew", padx=4, pady=4)
        ttk.Label(frm, text="Output File Path").grid(row=4, column=0, sticky="w", padx=4, pady=4)
        out_row = ttk.Frame(frm)
        out_row.grid(row=4, column=1, sticky="ew", padx=4, pady=4)
        ttk.Entry(out_row, textvariable=output_var, width=48).pack(side="left", fill="x", expand=True)

        def pick_output():
            p = filedialog.askopenfilename(
                title="Select existing output file to append to",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
            )
            if p:
                output_var.set(p)

        ttk.Button(out_row, text="Browse…", command=pick_output).pack(side="left", padx=(6, 0))

        def save():
            if not os.path.exists(_ENV_PATH):
                with open(_ENV_PATH, "w", encoding="utf-8"):
                    pass
            set_key(_ENV_PATH, "LYNX_URL", url_var.get().strip())
            set_key(_ENV_PATH, "LYNX_USER", user_var.get().strip())
            set_key(_ENV_PATH, "LYNX_PASSWORD", pw_var.get().strip())
            set_key(_ENV_PATH, "LYNX_WORK_FOLDER", work_var.get().strip() or "data")
            set_key(_ENV_PATH, "OUTPUT_FILE_PATH", output_var.get().strip())
            load_dotenv(_ENV_PATH, override=True)
            self.fetch_download_dir_var.set(self._env_get("LYNX_WORK_FOLDER", "data"))
            if output_var.get().strip():
                self.output_path_var.set(output_var.get().strip())
                self._refresh_merge_button_enabled()
            self._log("Fetch settings saved to .env")
            win.destroy()

        row = ttk.Frame(frm)
        row.grid(row=5, column=0, columnspan=2, sticky="e", pady=(8, 0))
        ttk.Button(row, text="Cancel", command=win.destroy).pack(side="right", padx=4)
        ttk.Button(row, text="Save", command=save).pack(side="right", padx=4)
        frm.columnconfigure(1, weight=1)

    # ── confirmation helpers ──────────────────────────────────────────────────

    def _show_confirm(self, msg):
        """Called from the main thread to display the confirm panel."""
        self.confirm_msg_var.set(msg)
        if self._merge_running:
            self.merge_status_var.set("Merge status: waiting for confirmation")
        self.frm_confirm.pack(fill="x", padx=10, pady=4)

    def _hide_confirm(self):
        self.frm_confirm.pack_forget()

    def _resolve_confirm(self, result: bool):
        """Called when user clicks Proceed or Abort."""
        self._hide_confirm()
        self._confirm_result = result
        self._confirm_event.set()

    def confirm(self, msg: str) -> bool:
        """
        Thread-safe blocking confirm. Called from the worker thread.
        Shows the prompt in the UI and waits for the user to click Proceed or Abort.
        """
        self._confirm_event.clear()
        self._confirm_result = False
        self.after(0, self._show_confirm, msg)
        self._confirm_event.wait()   # block worker thread
        return self._confirm_result

    # ── logging ───────────────────────────────────────────────────────────────

    def _log(self, msg):
        self.after(0, self._log_sync, msg)

    def _log_sync(self, msg):
        self.log_text.configure(state="normal")
        self.log_text.insert("end", msg + "\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def _clear_log(self):
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.configure(state="disabled")

    def _run_fetch(self):
        if self._fetch_running:
            return

        month = self._month_text(self.fetch_month_num_var.get(), self.fetch_year_var.get())
        try:
            month_date_range(month)
        except Exception as e:
            self._log(f"⚠  Invalid month: {e}")
            return

        load_dotenv(_ENV_PATH, override=True)
        url = self._env_get("LYNX_URL")
        user = self._env_get("LYNX_USER")
        pw = self._env_get("LYNX_PASSWORD")
        if not url or not user or not pw:
            self._log("⚠  Missing LYNX_URL / LYNX_USER / LYNX_PASSWORD. Open Settings first.")
            return

        download_dir = self.fetch_download_dir_var.get().strip() or self._env_get("LYNX_WORK_FOLDER", "data")
        if not download_dir:
            download_dir = "data"
        try:
            slow_mo = max(0, int(self.fetch_slow_seconds_var.get())) * 1000
        except Exception:
            self._log("⚠  Slow-mo seconds must be an integer.")
            return
        headless = bool(self.fetch_headless_var.get())

        self._fetch_running = True
        self._fetch_abort_requested = False
        self.fetch_btn.configure(state="disabled")
        self._set_action_button_state(self.fetch_btn, "running")
        self._refresh_lifecycle_state()
        self.fetch_status_var.set("Fetch status: running")
        self.fetch_result_name_var.set("Downloaded file: (running...)")
        self._log(f"Starting fetch | month={month} headless={headless} slow_mo={slow_mo} folder={download_dir}")

        def worker():
            try:
                with sync_playwright() as p:
                    browser = p.chromium.launch(headless=headless, slow_mo=slow_mo)
                    context = browser.new_context(accept_downloads=True)
                    self._fetch_browser = browser
                    self._fetch_context = context
                    page = context.new_page()
                    page.on("dialog", lambda dialog: dialog.accept())
                    try:
                        saved = download_superbill(
                            page,
                            download_dir,
                            month,
                            interactive=False,
                        )
                    finally:
                        context.close()
                        browser.close()
                        self._fetch_context = None
                        self._fetch_browser = None
                self.after(0, lambda: self._on_fetch_success(saved))
            except Exception as e:
                self.after(0, lambda: self._on_fetch_error(e))

        threading.Thread(target=worker, daemon=True).start()

    def _on_fetch_success(self, saved_path: str):
        self._fetch_running = False
        self.fetch_btn.configure(state="normal")
        self._set_action_button_state(self.fetch_btn, "success")
        self._refresh_lifecycle_state()
        self.input_path_var.set(saved_path)  # feed directly into merge stage
        self._refresh_merge_button_enabled()
        self.fetch_result_name_var.set(f"Downloaded file: {os.path.basename(saved_path)}")
        self.fetch_status_var.set("Fetch status: finished successfully")
        self._log(f"✓  Fetch completed: {saved_path}")
        self._log("✓  Input file auto-filled for merge stage.")

    def _on_fetch_error(self, err: Exception):
        self._fetch_running = False
        self.fetch_btn.configure(state="normal")
        if self._fetch_abort_requested:
            self.fetch_status_var.set("Fetch status: aborted")
            self._log("⚠  Fetch aborted by user.")
        else:
            self.fetch_status_var.set("Fetch status: failed")
            self._log(f"ERROR during fetch: {err}")
        self._set_action_button_state(self.fetch_btn, "failed")
        self._refresh_lifecycle_state()
        self.fetch_result_name_var.set("Downloaded file: (none)")

    # ── run ───────────────────────────────────────────────────────────────────

    def _run(self):
        input_path  = self.input_path_var.get().strip()
        output_path = self.output_path_var.get().strip()

        if not input_path:
            self._log("⚠  Please select an input file first.")
            return
        if not os.path.isfile(input_path):
            self._log(f"⚠  Input file not found: {input_path}")
            return
        if not output_path:
            self._log("⚠  Please specify an output file.")
            return
        if not os.path.isfile(output_path):
            self._log(f"⚠  Output file not found: {output_path}")
            return

        self._merge_running = True
        self._merge_abort_requested = False
        self.run_btn.configure(state="disabled")
        self._set_action_button_state(self.run_btn, "running")
        self._refresh_lifecycle_state()
        self.merge_status_var.set("Merge status: running")
        self._log("=" * 60)
        self._log("Starting merge…")

        def worker():
            try:
                ok = process(input_path, output_path, self._log, self.confirm)
                self.after(0, lambda: self._on_merge_done(ok))
            except Exception as e:
                tb = traceback.format_exc()
                self._log(f"ERROR during merge: {e}")
                self._log(tb)
                self.after(0, lambda: self._on_merge_error(e))
            finally:
                self.after(0, self._on_merge_finish_ui)

        threading.Thread(target=worker, daemon=True).start()

    def _on_merge_finish_ui(self):
        self._merge_running = False
        self._refresh_merge_button_enabled()
        self._refresh_lifecycle_state()

    def _on_merge_done(self, ok: bool):
        if ok:
            self._set_action_button_state(self.run_btn, "success")
            self.merge_status_var.set("Merge status: finished successfully")
            self._log("✓  Merge finished successfully.")
        else:
            self._set_action_button_state(self.run_btn, "failed")
            if self._merge_abort_requested:
                self.merge_status_var.set("Merge status: aborted")
                self._log("⚠  Merge aborted by user.")
            else:
                self.merge_status_var.set("Merge status: finished with failure/abort")
                self._log("⚠  Merge finished with failure or user abort.")

    def _on_merge_error(self, err: Exception):
        self._set_action_button_state(self.run_btn, "failed")
        self.merge_status_var.set(f"Merge status: error ({err})")


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    app = App()
    app.mainloop()
